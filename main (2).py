import openpyxl
from openpyxl.styles import PatternFill
wb=openpyxl.load_workbook("C:\\Users\\DELL\\OneDrive\\Desktop\\Sale.xlsx")
sh1=wb["Jan.21"]
sh2=wb["Feb.21"]
for i in range(2,9):
    v1=int(sh1.cell(i,10).value)
    v2=int(sh2.cell(i,10).value)
    if v2>v1:
        sh2.cell(i,10).fill=PatternFill("solid",fgColor="2dd22d")
    elif v2==v1:
        sh2.cell(i,10).fill=PatternFill("solid",fgColor="ecec13")
    else:
        sh2.cell(i,10).fill=PatternFill("solid",fgColor="d22d2d")
wb.save("C:\\Users\\DELL\\OneDrive\\Desktop\\Sale.xlsx")

